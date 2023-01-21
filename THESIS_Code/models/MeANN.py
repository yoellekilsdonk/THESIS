import merfadjustedANN
from hyperopt import hp, tpe, Trials, fmin, STATUS_OK
import numpy as np
import pandas as pd
import sklearn
import math
from keras.models import Sequential
from keras.layers import Dense, Dropout
from keras.callbacks import EarlyStopping, ModelCheckpoint
from tensorflow import keras
from keras.wrappers.scikit_learn import KerasRegressor
from datetime import datetime
import openpyxl

seed = 0
max_evals = 125

"""ANN model implementation. Can be tuned by Hyperopt."""


class MeKerasNN:
    def __init__(self, dataset, tune=False, normalization='none', params={}):
        self.dataset = dataset
        self.name = 'MeANN'
        self.normalization = normalization

        epoch = 100
        patience = 10
        norm = ['none','minmax' ,'quantile','standardize']
        act = ['relu', 'linear', 'sigmoid']
        nlayer = [1, 2]

        if not tune:
            self.params = params

            # Create necessary folds of dependent variable, fixed-effect regressors, random-effect regressors, and
            # clusters
            ytrainfold_finalfit = self.dataset.ytrainfold_finalfit[0]["hb"]  # Train
            ytestfold_finalfit = self.dataset.ytestfold_finalfit[0]["hb"]  # Test
            ytestfoldextra_finalfit = self.dataset.yextratest_finalfold[0]["hb"]  # Extra test

            Xtrainfold_finalfit = self.dataset.Xtrainfold_finalfit[0].drop(["id"], axis=1)
            Xtestfold_finalfit = self.dataset.Xtestfold_finalfit[0].drop(["id"], axis=1)
            Xtestfoldextra_finalfit = self.dataset.Xextratest_finalfold[0].drop(["id"], axis=1)

            Ztrainfinalfit = (pd.DataFrame(np.ones(np.size(Xtrainfold_finalfit, 0))))
            Ztestfinalfit = (pd.DataFrame(np.ones(np.size(Xtestfold_finalfit, 0))))
            Ztestextrafinalfit = (pd.DataFrame(np.ones(np.size(Xtestfoldextra_finalfit, 0))))

            clustertrainfinalfit = dataset.Xtrainfold_finalfit[0][['id']].squeeze()
            clustertestfinalfit = dataset.Xtestfold_finalfit[0][['id']].squeeze()
            clustertestextrafinalfit = dataset.Xextratest_finalfold[0][['id']].squeeze()

            # Create mixed-effects ANN model
            def baseline_modelfin():
                model = Sequential()
                model.add(Dense(params['n_neurons'], input_dim=Xtestfold_finalfit.shape[1],
                                activation=params['activation']))
                model.add(Dropout(params['init_dropout']))
                if params['n_layers'] == 2:
                    model.add(Dense(params['n_neurons'], activation=params['activation']))
                    model.add(Dropout(params['mid_dropout']))
                model.add(Dense(1, activation='relu'))

                opt = keras.optimizers.Adam(learning_rate=params['learning_rate'])
                model.compile(optimizer=opt, metrics=keras.metrics.RootMeanSquaredError(),
                              loss=keras.losses.MeanSquaredError())
                return model

            callbacks = [
                EarlyStopping(monitor='loss', patience=patience, mode='auto', verbose=0),
                ModelCheckpoint(
                    'MeANN_model_final', monitor='loss', mode='auto', save_best_only=True, verbose=0)
            ]

            model = KerasRegressor(  # Can't do self.model bc of pickling local object
                build_fn=baseline_modelfin,
                epochs=epoch,
                batch_size=pow(2, params['batch_size']),
                verbose=0,
                callbacks=callbacks)

            # Fit on training + validation set
            print("Fit final model: {}".format(datetime.now().isoformat(' ', 'seconds')))
            self.model = merfadjustedANN.MERF(model, max_iterations=5, currentrunname='MeANN_model_final')
            self.model.fit(self.dataset.normalize(Xtrainfold_finalfit, self.normalization),
                           Ztrainfinalfit, clustertrainfinalfit, ytrainfold_finalfit)

            # Predict on test set (containing observations for individuals we have never seen before)
            print("Predict final model: {}".format(datetime.now().isoformat(' ', 'seconds')))
            yhat = self.model.predict(self.dataset.normalize(Xtestfold_finalfit, self.normalization),
                                      Ztestfinalfit, clustertestfinalfit)
            yhat = yhat.reshape([-1, ])  # Compatibility with other models
            yhat[yhat < 0] = 0  # Cannot make negative predictions
            self.yhat = yhat
            self.rmse = math.sqrt(sklearn.metrics.mean_squared_error(ytestfold_finalfit, self.yhat))

            # Predict the extra test set (containing observations for individuals we have seen before)
            print("Predict extra analysis: {}".format(datetime.now().isoformat(' ', 'seconds')))
            yhat_extra = self.model.predict(self.dataset.normalize(Xtestfoldextra_finalfit, self.normalization),
                                            Ztestextrafinalfit, clustertestextrafinalfit)
            yhat_extra = yhat_extra.reshape([-1, ])  # Compatibility with other models
            yhat_extra[yhat_extra < 0] = 0  # Cannot make negative predictions
            self.yhat_extra = yhat_extra
            self.rmse_extra = math.sqrt(sklearn.metrics.mean_squared_error(ytestfoldextra_finalfit, self.yhat_extra))

        else:
            self.overview = {}

            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.cell(row=1, column=1).value = "Hyperopt iteration"
            sheet.cell(row=1, column=2).value = "n_neurons"
            sheet.cell(row=1, column=3).value = "learning_rate"
            sheet.cell(row=1, column=4).value = "batch_size"
            sheet.cell(row=1, column=5).value = "mid_dropout"
            sheet.cell(row=1, column=6).value = "activation"
            sheet.cell(row=1, column=7).value = "n_layers"
            sheet.cell(row=1, column=8).value = "init_dropout"
            sheet.cell(row=1, column=9).value = "normalization"
            sheet.cell(row=1, column=10).value = "Loss"
            wb.save('V:/UserData/079915/emcthesis/test.xlsx')
            #wb.save('/mnt/data/yoelle/test.xlsx')

            def objective(params):
                k = spark_trials.tids[-1] + 1  # Iteration of hyperopt
                self.overview[str(k)] = {}  # Create a dictionary of parameters per hyperopt iteration

                # Retreive parameters
                normalization = params['Normalization']

                ref_workbook = openpyxl.load_workbook('V:/UserData/079915/emcthesis/test.xlsx')
                #ref_workbook = openpyxl.load_workbook('/mnt/data/yoelle/test.xlsx')
                sheet = ref_workbook.active
                sheet.cell(row=k+1, column=9).value = params["Normalization"]

                params = {
                    'n_neurons': int(params['n_neurons']),
                    'learning_rate': float("{:.3f}".format(params['learning_rate'])),
                    'batch_size': int(params['batch_size']),
                    'mid_dropout': float("{:.3f}".format(params['mid_dropout'])),
                    'activation': params['activation'],
                    'n_layers': int(params['n_layers']),
                    'init_dropout': float("{:.3f}".format(params['init_dropout']))
                }

                rmse = []  # Create RMSE list (can be any other metric)
                for i in range(0, 1):
                    # Create necessary folds of dependent variable and regressors
                    ytrainfold = dataset.ytrainfolds[i]["hb"]  # Train
                    yvalidationfold = dataset.yvalidationfolds[i]["hb"]  # Validation

                    Xtrainfold = dataset.Xtrainfolds[i].drop(["id"], axis=1)
                    Xvalidationfold = dataset.Xvalidationfolds[i].drop(["id"], axis=1)

                    Ztrainfold = (pd.DataFrame(np.ones(np.size(dataset.Xtrainfolds[i], 0))))
                    Zvalidation = (pd.DataFrame(np.ones(np.size(Xvalidationfold, 0))))

                    clustervalidationfold = dataset.Xvalidationfolds[i][['id']].squeeze()
                    clustertrainfold = dataset.Xtrainfolds[i][['id']].squeeze()

                    # Create mixed-effects ANN tune model
                    def baseline_model():
                        model = Sequential()
                        model.add(Dense(params['n_neurons'], input_dim=Xtrainfold.shape[1],
                                        activation=params['activation']))
                        model.add(Dropout(params['init_dropout']))
                        if params['n_layers'] == 2:
                            model.add(Dense(params['n_neurons'], activation=params['activation']))
                            model.add(Dropout(params['mid_dropout']))
                        model.add(Dense(1, activation='relu'))

                        opt = keras.optimizers.Adam(learning_rate=params['learning_rate'])
                        model.compile(optimizer=opt, metrics=keras.metrics.RootMeanSquaredError(),
                                      loss=keras.losses.MeanSquaredError())
                        return model

                    callbacks = [
                        EarlyStopping(monitor='loss', patience=patience, mode='auto', verbose=0),
                        ModelCheckpoint(
                            'MeANN_model_tune', monitor='loss', mode='auto', save_best_only=True, verbose=0)
                    ]

                    nnet = KerasRegressor(
                        build_fn=baseline_model,
                        epochs=epoch,
                        batch_size=pow(2, params['batch_size']),
                        verbose=0,
                        callbacks=callbacks)

                    # Train on test set
                    print("Fit fold {}: {}".format(i, datetime.now().isoformat(' ', 'seconds')))
                    meANN = merfadjustedANN.MERF(nnet, max_iterations=5, currentrunname='MeANN_model_tune')
                    meANN.fit(self.dataset.normalize(Xtrainfold, normalization), Ztrainfold, clustertrainfold,
                              ytrainfold)

                    # Predict on validation set
                    print("Predict fold {}: {}".format(i, datetime.now().isoformat(' ', 'seconds')))
                    yhat = meANN.predict(self.dataset.normalize(Xvalidationfold, normalization), Zvalidation,
                                         clustervalidationfold)
                    yhat = yhat.reshape([-1, ])  # Compatibility with other models
                    yhat[yhat < 0] = 0
                    rmse.append(sklearn.metrics.mean_squared_error(yvalidationfold, yhat))
                    print(rmse)
                    print("Done fold {}: {}".format(i, datetime.now().isoformat(' ', 'seconds')))

                    sheet.cell(row=k + 1, column=1).value = k
                    sheet.cell(row=k + 1, column=2).value = params['n_neurons']
                    sheet.cell(row=k + 1, column=3).value = params["learning_rate"]
                    sheet.cell(row=k + 1, column=4).value = params["batch_size"]
                    sheet.cell(row=k + 1, column=5).value = params["mid_dropout"]
                    sheet.cell(row=k + 1, column=6).value = params["activation"]
                    sheet.cell(row=k + 1, column=7).value = params["n_layers"]
                    sheet.cell(row=k + 1, column=8).value = params["init_dropout"]
                    sheet.cell(row=k + 1, column=10).value = rmse[-1]
                    ref_workbook.save('V:/UserData/079915/emcthesis/test.xlsx')
                    #ref_workbook.save('/mnt/data/yoelle/test.xlsx')

                    # Save parameters, rmse, and summary statistics of current hyperopt run
                    self.overview[str(k)][str(i)] = {}
                    self.overview[str(k)][str(i)]['params'] = params
                    self.overview[str(k)][str(i)]['rmse'] = sklearn.metrics.mean_squared_error(yvalidationfold, yhat)
                    self.overview[str(k)][str(i)]['yhat'] = pd.DataFrame(yhat).describe()
                return {'loss': np.mean(rmse), 'status': STATUS_OK}

            # Define search space for hyperopt
            search_space = {
                'type': 'MeANN',
                'n_neurons': hp.quniform('n_neurons', 8, 240, 4),  # 8 is 2/3 the input dim of X
                'learning_rate': hp.loguniform('learning_rate', np.log(0.0001), np.log(1)),
                'batch_size': hp.uniform('batch_size', 11, 15),
                'init_dropout': hp.uniform('init_dropout', 0.0, 1.0),  # Dropout in first layer
                'mid_dropout': hp.uniform('mid_dropout', 0.0, 1.0),  # Dropout in second layer
                'activation': hp.choice('activation', act),  # ['relu','linear','sigmoid']
                'n_layers': hp.choice('n_layers', nlayer),  # [1,2]
                'Normalization': hp.choice('Normalization', norm)  # ['none', 'minmax', 'quantile', 'standardize']
            }

            # Start hyperopt
            algo = tpe.suggest
            spark_trials = Trials()
            rstate = np.random.default_rng(self.dataset.seed)  # <== Use any number here but fixed

            best_result = fmin(
                fn=objective,
                space=search_space,
                algo=algo,
                max_evals=max_evals,
                trials=spark_trials,
                rstate=rstate)

            # Save and print best hyperparameter results
            best_result['n_neurons'] = int(best_result['n_neurons'])
            best_result['batch_size'] = int(best_result['batch_size'])  # 2 ^ this number
            best_result['activation'] = act[best_result['activation']]
            best_result['n_layers'] = nlayer[best_result['n_layers']]
            self.normalization = norm[best_result['Normalization']]
            print(best_result)
            self.params = best_result
            self.params.pop('Normalization')
